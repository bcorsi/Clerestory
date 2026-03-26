#!/usr/bin/env python3
"""
Clerestory — Underwriting Model Generator
Generates a branded 4-sheet Excel workbook from deal inputs.

Usage:
    python3 generate_uw_model.py '<json_inputs>' [output_path]

Outputs the resolved output path to stdout on success.
"""
import sys
import json
import math
import os

# ── Input parsing ─────────────────────────────────────────────────────────────

def parse_num(val, strip='$,'):
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val)
    for c in strip:
        s = s.replace(c, '')
    try:
        return float(s.strip())
    except ValueError:
        return 0.0

# ── Finance math (mirrors DealDetail.jsx runModel) ───────────────────────────

def approximate_irr(cfs, lo=-0.5, hi=5.0, iters=120):
    for _ in range(iters):
        mid = (lo + hi) / 2
        npv = sum(cf / (1 + mid) ** t for t, cf in enumerate(cfs))
        if abs(npv) < 0.01:
            return mid
        if npv > 0:
            lo = mid
        else:
            hi = mid
    return (lo + hi) / 2

def run_model(inputs):
    price    = parse_num(inputs.get('price',    '47500000'))
    rent     = parse_num(inputs.get('rent',     '0.82'))
    sf       = parse_num(inputs.get('sf',       '312000'))
    exit_cap = parse_num(inputs.get('exitCap',  '5.25')) / 100
    ltv      = parse_num(inputs.get('ltv',      '65'))   / 100
    rate     = parse_num(inputs.get('rate',     '6.50')) / 100
    bumps    = parse_num(inputs.get('bumps',    '3.0'))  / 100
    hold     = max(1, int(parse_num(inputs.get('hold', '5'))))
    mkt_rent = parse_num(inputs.get('marketRent', '0.98'))

    noi0     = rent * sf * 12
    going_in = (noi0 / price * 100) if price > 0 else 0

    # Year-by-year NOI (mirrors JS: noi0 * (1+bumps)^(y-1), y=1..hold)
    noi_arr = [noi0 * (1 + bumps) ** y for y in range(hold)]

    # Debt
    loan     = price * ltv
    equity   = price * (1 - ltv)
    mo_rate  = rate / 12
    mo_pmt   = (loan * mo_rate / (1 - (1 + mo_rate) ** -360)) if rate > 0 else 0
    ds       = mo_pmt * 12
    dscr     = (noi_arr[0] / ds) if ds > 0 else 0

    # Exit
    exit_noi = noi0 * (1 + bumps) ** hold
    exit_val = (exit_noi / exit_cap) if exit_cap > 0 else 0

    # Cash flows
    cf_unlev = [-price] + noi_arr[:-1] + [noi_arr[-1] + exit_val]
    cf_lev   = [-(price - loan)] \
             + [n - ds for n in noi_arr[:-1]] \
             + [noi_arr[-1] - ds + exit_val - loan]

    unlev_irr = approximate_irr(cf_unlev)
    lev_irr   = approximate_irr(cf_lev)

    # Equity multiple (mirrors JS formula)
    em = (sum(cf_lev[1:]) + equity) / equity if equity > 0 else 0

    return dict(
        price=price, rent=rent, sf=sf, exit_cap=exit_cap, ltv=ltv,
        rate=rate, bumps=bumps, hold=hold, mkt_rent=mkt_rent,
        noi0=noi0, going_in=going_in, noi_arr=noi_arr,
        loan=loan, equity=equity, ds=ds, dscr=dscr,
        exit_val=exit_val, exit_noi=exit_noi,
        cf_unlev=cf_unlev, cf_lev=cf_lev,
        unlev_irr=unlev_irr, lev_irr=lev_irr, em=em,
    )

# ── Excel generation ──────────────────────────────────────────────────────────

def build_excel(inputs, r, output_path):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        import subprocess
        subprocess.check_call(
            [sys.executable, '-m', 'pip', 'install', 'openpyxl', '-q'],
            stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
        )
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

    # ── Color palette (matches Clerestory design system) ──
    NAVY    = '3B5F8A'
    NAVY_LT = 'D6E4F4'
    SLATE   = '6480A2'
    OFF_WHT = 'F4F1EC'
    WHITE   = 'FFFFFF'
    DARK    = '1A2433'
    MID     = '4A5568'
    LGRAY   = 'E2E8F0'
    GREEN   = '156636'
    GRN_LT  = 'E8F5EE'
    AMBER   = '8C5A04'
    AMB_LT  = 'FDF5E8'
    RUST    = '8B2500'
    RST_LT  = 'FEECE8'

    def f(size=10, bold=False, color=DARK, italic=False):
        return Font(name='Calibri', size=size, bold=bold, color=color, italic=italic)

    def fill(color):
        return PatternFill('solid', fgColor=color)

    def center():
        return Alignment(horizontal='center', vertical='center')

    def right():
        return Alignment(horizontal='right', vertical='center')

    def thin_border(sides='bottom'):
        s = Side(style='thin', color='C0C8D2')
        b = Border()
        if 'bottom' in sides: b.bottom = s
        if 'top'    in sides: b.top    = s
        if 'left'   in sides: b.left   = s
        if 'right'  in sides: b.right  = s
        return b

    def w(ws, col, width):
        ws.column_dimensions[col if isinstance(col, str) else get_column_letter(col)].width = width

    def h(ws, row, height):
        ws.row_dimensions[row].height = height

    hold      = r['hold']
    deal_name = inputs.get('dealName', 'Deal')
    address   = inputs.get('address', '')
    heading   = deal_name + (f' — {address}' if address else '')

    wb = Workbook()

    # ══════════════════════════════════════════════════════════════
    # SHEET 1  ·  RETURNS SUMMARY
    # ══════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = 'Returns Summary'
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 100

    # Title block
    ws.merge_cells('A1:F1')
    ws['A1'] = 'CLERESTORY  ·  Underwriting Model'
    ws['A1'].font = f(16, bold=True, color=NAVY)
    ws['A1'].alignment = Alignment(vertical='center')
    h(ws, 1, 28)

    ws.merge_cells('A2:F2')
    ws['A2'] = heading
    ws['A2'].font = f(11, bold=True, color=DARK)
    h(ws, 2, 18)

    ws.merge_cells('A3:F3')
    ws['A3'] = f'{hold}-Year Hold Period  ·  Generated by Clerestory'
    ws['A3'].font = f(9, color=MID, italic=True)
    h(ws, 3, 14)

    h(ws, 4, 8)

    # KPI header
    for col, label in enumerate(['Metric', 'Value', 'Benchmark', 'Status', 'Notes', ''], 1):
        c = ws.cell(5, col, label)
        c.font = f(10, bold=True, color=WHITE)
        c.fill = fill(NAVY)
        c.alignment = center()
    h(ws, 5, 22)

    def irr_status(val):
        if val > 0.16: return ('✓ Strong', GREEN, GRN_LT)
        if val > 0.12: return ('✓ Meets Hurdle', GREEN, GRN_LT)
        return ('⚠ Below Hurdle', RUST, RST_LT)

    kpis = [
        ('Going-In Cap Rate',  f"{r['going_in']:.2f}%",          'Market: 5.0–5.5%',
         '—', MID, OFF_WHT, 'Initial yield on purchase price'),
        ('Unlevered IRR',      f"{r['unlev_irr']*100:.1f}%",     'Hurdle: 10%',
         *irr_status(r['unlev_irr'])[:2], irr_status(r['unlev_irr'])[2], 'Before leverage'),
        ('Levered IRR',        f"{r['lev_irr']*100:.1f}%",       'Hurdle: 12–16%',
         *irr_status(r['lev_irr'])[:2], irr_status(r['lev_irr'])[2], 'After leverage & fees'),
        ('Equity Multiple',    f"{r['em']:.2f}×",                'Target: > 1.5×',
         ('✓ Strong' if r['em'] > 1.5 else '✓ OK' if r['em'] > 1.2 else '⚠ Low'),
         (GREEN if r['em'] > 1.2 else AMBER),
         (GRN_LT if r['em'] > 1.2 else AMB_LT),
         'Total return / equity invested'),
        ('DSCR Year 1',        f"{r['dscr']:.2f}×",              'Min: 1.20×',
         ('✓ Meets Min' if r['dscr'] > 1.2 else '✗ Below Min'),
         (GREEN if r['dscr'] > 1.2 else RUST),
         (GRN_LT if r['dscr'] > 1.2 else RST_LT),
         'Debt Service Coverage Ratio'),
        ('Exit Cap Rate',      f"{r['exit_cap']*100:.2f}%",      'Assumption',
         '—', MID, OFF_WHT, 'Applied to final year NOI'),
        ('Annual Rent Bumps',  f"{r['bumps']*100:.1f}%",         'Assumption',
         '—', MID, OFF_WHT, 'Compounding annual escalation'),
    ]

    for i, (metric, val, bench, status, st_color, row_fill, note) in enumerate(kpis):
        row_n = 6 + i
        bg = row_fill if i % 2 == 0 else WHITE
        for col in range(1, 7):
            ws.cell(row_n, col).fill = fill(bg)

        ws.cell(row_n, 1, metric).font      = f(10, bold=True, color=DARK)
        ws.cell(row_n, 2, val).font         = f(13, bold=True, color=NAVY)
        ws.cell(row_n, 2).alignment         = center()
        ws.cell(row_n, 3, bench).font       = f(9, color=MID, italic=True)
        ws.cell(row_n, 3).alignment         = center()
        ws.cell(row_n, 4, status).font      = f(10, bold=True, color=st_color)
        ws.cell(row_n, 4).alignment         = center()
        ws.cell(row_n, 5, note).font        = f(9, color=MID, italic=True)
        h(ws, row_n, 20)

    for col, width in [('A', 22), ('B', 14), ('C', 20), ('D', 18), ('E', 32), ('F', 4)]:
        w(ws, col, width)

    # ══════════════════════════════════════════════════════════════
    # SHEET 2  ·  INPUTS
    # ══════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet('Inputs')
    ws2.sheet_view.showGridLines = False

    ws2.merge_cells('A1:C1')
    ws2['A1'] = 'Deal Inputs'
    ws2['A1'].font = f(14, bold=True, color=NAVY)
    h(ws2, 1, 24)

    ws2.merge_cells('A2:C2')
    ws2['A2'] = heading
    ws2['A2'].font = f(10, color=MID, italic=True)
    h(ws2, 2, 16)
    h(ws2, 3, 8)

    rows = [
        ('Purchase Price',            f"${r['price']:,.0f}",             'deal_input'),
        ('Building SF',               f"{r['sf']:,.0f} SF",              'deal_input'),
        ('In-Place Rent (NNN/SF/Mo)', f"${r['rent']:.2f}",               'deal_input'),
        ('Market Rent (NNN/SF/Mo)',   f"${r['mkt_rent']:.2f}",           'deal_input'),
        ('Hold Period',               f"{r['hold']} Years",               'assumption'),
        ('', '', ''),
        ('Exit Cap Rate',             f"{r['exit_cap']*100:.2f}%",       'assumption'),
        ('Annual Rent Bumps',         f"{r['bumps']*100:.1f}%",          'assumption'),
        ('LTV',                       f"{r['ltv']*100:.0f}%",            'assumption'),
        ('Interest Rate',             f"{r['rate']*100:.2f}%",           'assumption'),
        ('', '', ''),
        ('Loan Amount',               f"${r['loan']:,.0f}",              'derived'),
        ('Equity Required',           f"${r['equity']:,.0f}",            'derived'),
        ('Annual Debt Service',       f"${r['ds']:,.0f}",                'derived'),
        ('Year 1 NOI',                f"${r['noi0']:,.0f}",              'derived'),
        ('Going-In Cap Rate',         f"{r['going_in']:.2f}%",           'derived'),
        ('Exit NOI (Year {})'.format(r['hold']+1),
                                      f"${r['exit_noi']:,.0f}",          'derived'),
        ('Exit Value',                f"${r['exit_val']:,.0f}",          'derived'),
    ]

    section_colors = {'deal_input': NAVY_LT, 'assumption': GRN_LT, 'derived': OFF_WHT}
    label_colors   = {'deal_input': NAVY,     'assumption': GREEN,   'derived': MID}

    row_i = 4
    for label, value, kind in rows:
        if not label:
            h(ws2, row_i, 8)
            row_i += 1
            continue
        bg = section_colors.get(kind, WHITE)
        lc = label_colors.get(kind, DARK)
        ws2.cell(row_i, 1, label).font = f(10, color=DARK)
        ws2.cell(row_i, 1).fill = fill(bg)
        ws2.cell(row_i, 2, value).font = f(11, bold=True, color=lc)
        ws2.cell(row_i, 2).fill = fill(bg)
        ws2.cell(row_i, 2).alignment = right()
        ws2.cell(row_i, 3).fill = fill(bg)
        h(ws2, row_i, 18)
        row_i += 1

    # Legend
    row_i += 1
    for kind, label in [('deal_input', 'Deal Input'), ('assumption', 'Assumption'), ('derived', 'Derived / Calculated')]:
        ws2.cell(row_i, 1).fill = fill(section_colors[kind])
        ws2.cell(row_i, 2, label).font = f(9, color=MID, italic=True)
        row_i += 1

    for col, width in [('A', 30), ('B', 18), ('C', 4)]:
        w(ws2, col, width)

    # ══════════════════════════════════════════════════════════════
    # SHEET 3  ·  CASH FLOWS
    # ══════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet('Cash Flows')
    ws3.sheet_view.showGridLines = False

    ws3.merge_cells(f'A1:{get_column_letter(hold+3)}1')
    ws3['A1'] = 'Annual Cash Flow Model'
    ws3['A1'].font = f(14, bold=True, color=NAVY)
    h(ws3, 1, 24)

    ws3.merge_cells(f'A2:{get_column_letter(hold+3)}2')
    ws3['A2'] = heading
    ws3['A2'].font = f(10, color=MID, italic=True)
    h(ws3, 2, 16)
    h(ws3, 3, 8)

    # Column headers
    headers = [''] + ['Acquisition'] + [f'Year {y+1}' for y in range(hold)] + ['Total']
    for col, hdr in enumerate(headers, 1):
        c = ws3.cell(4, col, hdr)
        c.font = f(10, bold=True, color=WHITE)
        c.fill = fill(NAVY)
        c.alignment = center()
        w(ws3, col, 14)
    w(ws3, 'A', 26)
    h(ws3, 4, 20)

    noi_arr   = r['noi_arr']
    exit_val  = r['exit_val']
    ds_annual = r['ds']
    loan      = r['loan']
    price     = r['price']
    cf_unlev  = r['cf_unlev']
    cf_lev    = r['cf_lev']

    def fmt_cell(ws, row, col, value, bold=False, is_neg=None):
        if value is None:
            return
        negative = (is_neg if is_neg is not None else value < 0)
        color = RUST if negative else (GREEN if value > 0 else MID)
        c = ws.cell(row, col, value)
        c.font = f(10, bold=bold, color=color)
        c.number_format = '#,##0'
        c.alignment = right()

    cf_sections = [
        # label, acq, [year values], total, bold
        ('— UNLEVERED —', None, [None]*hold, None, False),
        ('NOI',           None, noi_arr,           sum(noi_arr), False),
        ('Exit Value',    None, [None]*(hold-1)+[exit_val], exit_val, False),
        ('Total Unlevered CF', cf_unlev[0], cf_unlev[1:], sum(cf_unlev), True),
        ('', None, [None]*hold, None, False),
        ('— LEVERED —',   None, [None]*hold, None, False),
        ('Loan Proceeds', loan, [None]*hold, loan, False),
        ('Annual Debt Service', None, [-ds_annual]*hold, -ds_annual*hold, False),
        ('Loan Repayment', None, [None]*(hold-1)+[-loan], -loan, False),
        ('Total Levered CF', cf_lev[0], cf_lev[1:], sum(cf_lev), True),
    ]

    for i, (label, acq, yr_vals, total, bold) in enumerate(cf_sections):
        row_n = 5 + i
        bg = LGRAY if bold else (OFF_WHT if i % 2 == 0 else WHITE)
        if not label:
            h(ws3, row_n, 8)
            continue
        if label.startswith('—'):
            ws3.cell(row_n, 1, label).font = f(9, bold=True, color=SLATE, italic=True)
            h(ws3, row_n, 14)
            continue

        for col in range(1, hold + 4):
            ws3.cell(row_n, col).fill = fill(bg)
        ws3.cell(row_n, 1, label).font = f(10, bold=bold, color=DARK)
        h(ws3, row_n, 18)

        if acq is not None:
            fmt_cell(ws3, row_n, 2, acq, bold=bold)
        for yi, v in enumerate(yr_vals):
            if v is not None:
                fmt_cell(ws3, row_n, yi + 3, v, bold=bold)
        if total is not None:
            fmt_cell(ws3, row_n, hold + 3, total, bold=True)

    # ══════════════════════════════════════════════════════════════
    # SHEET 4  ·  SENSITIVITY
    # ══════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet('Sensitivity')
    ws4.sheet_view.showGridLines = False

    ws4.merge_cells('A1:G1')
    ws4['A1'] = 'IRR Sensitivity — Exit Cap Rate vs. Rent Growth'
    ws4['A1'].font = f(14, bold=True, color=NAVY)
    h(ws4, 1, 24)

    ws4.merge_cells('A2:G2')
    ws4['A2'] = 'Levered IRR  ·  Base case (5.25% cap / 3.0% bumps) highlighted in blue'
    ws4['A2'].font = f(9, color=MID, italic=True)
    h(ws4, 2, 14)
    h(ws4, 3, 8)

    exit_caps    = [4.50, 4.75, 5.00, 5.25, 5.50, 5.75]
    rent_growths = [2.0, 2.5, 3.0, 3.5, 4.0]

    # Header row
    ws4.cell(4, 1, 'Exit Cap \\ Rent Growth').font = f(9, bold=True, color=WHITE)
    ws4.cell(4, 1).fill = fill(NAVY)
    ws4.cell(4, 1).alignment = center()
    for j, rg in enumerate(rent_growths):
        c = ws4.cell(4, j + 2, f'{rg:.1f}%')
        c.font = f(10, bold=True, color=WHITE)
        c.fill = fill(NAVY)
        c.alignment = center()
    h(ws4, 4, 20)

    for i, cap in enumerate(exit_caps):
        row_n = 5 + i
        cap_cell = ws4.cell(row_n, 1, f'{cap:.2f}%')
        cap_cell.font = f(10, bold=True, color=SLATE)
        cap_cell.fill = fill(OFF_WHT)
        cap_cell.alignment = center()
        h(ws4, row_n, 20)

        for j, rg in enumerate(rent_growths):
            test_res = run_model(dict(inputs, exitCap=str(cap), bumps=str(rg)))
            irr = test_res['lev_irr']
            is_base = (abs(cap - 5.25) < 0.01 and abs(rg - 3.0) < 0.01)

            txt = f'{irr * 100:.1f}%'
            if is_base:
                bg, color, bold = NAVY_LT, NAVY, True
            elif irr > 0.16:
                bg, color, bold = GRN_LT,  GREEN, False
            elif irr > 0.12:
                bg, color, bold = OFF_WHT,  SLATE, False
            else:
                bg, color, bold = AMB_LT,  AMBER, False

            c = ws4.cell(row_n, j + 2, txt)
            c.font = f(11, bold=bold, color=color)
            c.fill = fill(bg)
            c.alignment = center()

    for col, width in [('A', 22)] + [(get_column_letter(j + 2), 12) for j in range(5)]:
        w(ws4, col, width)

    # Legend below grid
    legend_row = 5 + len(exit_caps) + 2
    for code, label in [(GRN_LT, '≥ 16% IRR — Strong'), (OFF_WHT, '12–16% IRR — Meets Hurdle'), (AMB_LT, '< 12% IRR — Below Hurdle'), (NAVY_LT, 'Base Case')]:
        ws4.cell(legend_row, 1).fill = fill(code)
        ws4.cell(legend_row, 1, '  ').font = f(10)
        ws4.cell(legend_row, 2, label).font = f(9, color=MID, italic=True)
        legend_row += 1

    # ── Save ──────────────────────────────────────────────────────
    wb.save(output_path)
    return output_path


# ── Entry point ───────────────────────────────────────────────────────────────

if __name__ == '__main__':
    if len(sys.argv) < 2:
        print('Usage: python3 generate_uw_model.py \'<json_inputs>\' [output_path]', file=sys.stderr)
        sys.exit(1)

    try:
        inputs = json.loads(sys.argv[1])
    except json.JSONDecodeError as e:
        print(f'Invalid JSON input: {e}', file=sys.stderr)
        sys.exit(1)

    output_path = sys.argv[2] if len(sys.argv) > 2 else f'/tmp/uw_model_{os.getpid()}.xlsx'

    results = run_model(inputs)
    path = build_excel(inputs, results, output_path)
    print(path, end='')  # stdout — consumed by route.js
